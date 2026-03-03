from __future__ import annotations

import io
import json
from urllib.parse import urlparse
from typing import Any

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side,
)
from openpyxl.utils import get_column_letter
from slugify import slugify

# ── Column definitions (matches Article_Writing_Winsome.xlsx Sheet1 exactly) ────

COLUMNS = [
    "Images :",
    "Recipe :",
    "Recipe Title :",
    "SEO title :",
    "Meta Desc :",
    "Categories :",
    "Article :",
    "URL Slug :",
    "Keyphrase :",
    "Tags :",
    "Pin Title :",
    "Pin Description :",
    "Keywords :",
    "Board :",
    "JSON :",
]

# Column widths (approximate)
COL_WIDTHS = [50, 60, 45, 50, 55, 18, 60, 45, 35, 45, 45, 55, 45, 30, 60]

_HEADER_FILL = PatternFill("solid", fgColor="2D5016")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL    = PatternFill("solid", fgColor="F5F0E8")
_BORDER_SIDE = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE,  bottom=_BORDER_SIDE,
)
_WRAP        = Alignment(wrap_text=True, vertical="top")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _images_newline(recipe: Any) -> str:
    """Return image URLs newline-separated (matches Article_Writing_Winsome)."""
    out: list[str] = []
    if recipe.generated_images:
        try:
            imgs = json.loads(recipe.generated_images)
            if imgs and isinstance(imgs, list):
                out = [str(u).strip() for u in imgs if u]
        except Exception:
            pass
    if not out and recipe.image_url:
        out = [recipe.image_url]
    return "\n".join(out) if out else ""


def _recipe_title(recipe: Any) -> str:
    """Recipe title from JSON name, recipe_text first line, or focus_keyword."""
    if recipe.generated_json:
        try:
            data = json.loads(recipe.generated_json)
            if isinstance(data, dict) and data.get("name"):
                return str(data["name"]).strip()
        except Exception:
            pass
    if recipe.recipe_text:
        first = recipe.recipe_text.split("\n")[0].strip()
        if first:
            return first
    return recipe.focus_keyword or ""


def _url_slug(recipe: Any) -> str:
    """URL slug from wp_permalink or slugify of title."""
    if recipe.wp_permalink:
        path = urlparse(recipe.wp_permalink).path.strip("/")
        if path:
            return path.split("/")[-1] or ""
    return slugify(_recipe_title(recipe)) or ""


def _recipe_row(recipe: Any) -> list[str]:
    title = _recipe_title(recipe)
    pin_title = recipe.focus_keyword or title
    return [
        _images_newline(recipe),                  # Images :
        recipe.generated_full_recipe or "",       # Recipe :
        title,                                    # Recipe Title :
        title,                                    # SEO title : (same as recipe title)
        recipe.meta_description or "",             # Meta Desc :
        recipe.category or "",                    # Categories :
        recipe.generated_article or "",           # Article :
        _url_slug(recipe),                        # URL Slug :
        recipe.focus_keyword or "",               # Keyphrase :
        "",                                       # Tags :
        pin_title,                                # Pin Title :
        recipe.meta_description or "",            # Pin Description :
        recipe.focus_keyword or "",               # Keywords :
        "",                                       # Board :
        recipe.generated_json or "",              # JSON :
    ]


# ── Sheet builder ─────────────────────────────────────────────────────────────

def _write_sheet(ws, site: Any, recipes: list[Any]) -> None:
    # Header row
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Set column widths
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, recipe in enumerate(recipes, start=2):
        row_data = _recipe_row(recipe)
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = _CELL_BORDER
            cell.alignment = _WRAP
            if fill:
                cell.fill = fill

    # Freeze header row
    ws.freeze_panes = "A2"


# ── Public API ────────────────────────────────────────────────────────────────

def build_project_excel(
    project_name: str,
    sites_with_recipes: list[tuple[Any, list[Any]]],
) -> bytes:
    """
    Build an xlsx workbook with one sheet per site.
    sites_with_recipes: list of (Site, [Recipe, ...])
    Returns bytes ready to stream.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for site, recipes in sites_with_recipes:
        domain = site.domain.replace("https://", "").replace("http://", "").rstrip("/")
        ws = wb.create_sheet(title=domain[:31])  # Excel max 31 chars
        _write_sheet(ws, site, recipes)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_site_excel(site: Any, recipes: list[Any]) -> bytes:
    """Build an xlsx workbook for a single site."""
    return build_project_excel(site.domain, [(site, recipes)])
